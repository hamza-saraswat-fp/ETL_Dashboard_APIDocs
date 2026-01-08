"""
Silver to Excel Loader - Main loader module
Converts Silver JSON to formatted Excel workbook
"""
import json
import os
from pathlib import Path
from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

from .excel_formatter import ExcelFormatter
from .config import (
    EXCEL_COLUMNS,
    EXCEL_COLUMN_DESCRIPTIONS,
    CUSTOM_FILTERS,
    FILTER_COLUMNS,
    FILTER_HEADER
)


class SilverToExcelLoader:
    """Loads silver JSON and generates Excel workbook"""

    def __init__(self, costbook_title: str = "WinSupply"):
        """
        Initialize loader

        Args:
            costbook_title: Title for the costbook
        """
        self.costbook_title = costbook_title
        self.formatter = ExcelFormatter(costbook_title)

    def load_silver_json(self, input_path: str) -> Dict:
        """Load silver JSON file"""
        with open(input_path, 'r') as f:
            return json.load(f)

    def process_systems(self, silver_data: Dict) -> List[Dict]:
        """
        Process all systems and generate row data

        Args:
            silver_data: Silver JSON data

        Returns:
            List of row dicts for the main sheet
        """
        systems = silver_data.get('systems', [])
        all_rows = []

        for system in systems:
            try:
                rows = self.formatter.format_system(system)
                all_rows.extend(rows)
            except Exception as e:
                print(f"Warning: Error formatting system {system.get('system_id', 'unknown')}: {e}")
                continue

        return all_rows

    def create_excel(self, rows: List[Dict], output_path: str):
        """
        Create Excel workbook with two sheets

        Args:
            rows: List of row dicts
            output_path: Path to output Excel file
        """
        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # Create writer object
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Create filters sheet first
            self._create_filters_sheet(writer)

            # Create main sheet
            self._create_main_sheet(rows, writer)

        # Apply additional formatting
        self._apply_formatting(output_path)

        print(f"Excel file created: {output_path}")
        print(f"Total rows: {len(rows)}")
        print(f"Total systems: {len(set(row['Job Name'] for row in rows))}")

    def _create_main_sheet(self, rows: List[Dict], writer: pd.ExcelWriter):
        """Create the main 'Flatrate Jobs & Catagories' sheet"""

        # Create header row with descriptions
        header_df = pd.DataFrame([EXCEL_COLUMN_DESCRIPTIONS], columns=EXCEL_COLUMNS)

        # Create column names row
        columns_df = pd.DataFrame([EXCEL_COLUMNS], columns=EXCEL_COLUMNS)

        # Create filter labels row (dashes for standard columns, filter names for custom filters)
        filter_labels = ['-'] * 20  # Dashes for columns A-T (indices 0-19)
        # Add filter names from CUSTOM_FILTERS for columns U onwards
        for filter_def in CUSTOM_FILTERS:
            filter_labels.append(filter_def['name'])
        # Pad with empty strings if we have fewer than 12 custom filters
        while len(filter_labels) < len(EXCEL_COLUMNS):
            filter_labels.append('')

        filter_labels_df = pd.DataFrame([filter_labels], columns=EXCEL_COLUMNS)

        # Create data rows
        if rows:
            data_df = pd.DataFrame(rows, columns=EXCEL_COLUMNS)
        else:
            data_df = pd.DataFrame(columns=EXCEL_COLUMNS)

        # Combine all rows: descriptions, column names, filter labels, then data
        combined_df = pd.concat([header_df, columns_df, filter_labels_df, data_df], ignore_index=True)

        # Write to Excel
        combined_df.to_excel(
            writer,
            sheet_name='Flatrate Jobs & Catagories',
            index=False,
            header=False
        )

    def _create_filters_sheet(self, writer: pd.ExcelWriter):
        """Create the 'Filters ' sheet"""

        # Create header row with descriptions
        header_row = pd.DataFrame([FILTER_COLUMNS], columns=FILTER_HEADER)

        # Create column names row
        columns_row = pd.DataFrame([FILTER_HEADER], columns=FILTER_HEADER)

        # Create filter data rows
        filter_data = []
        for filter_def in CUSTOM_FILTERS:
            filter_data.append({
                "Filter Name": filter_def["name"],
                "Filter Type": filter_def["type"]
            })

        if filter_data:
            data_df = pd.DataFrame(filter_data, columns=FILTER_HEADER)
        else:
            data_df = pd.DataFrame(columns=FILTER_HEADER)

        # Combine all rows
        combined_df = pd.concat([header_row, columns_row, data_df], ignore_index=True)

        # Write to Excel (note the space after 'Filters')
        combined_df.to_excel(
            writer,
            sheet_name='Filters ',
            index=False,
            header=False
        )

    def _apply_formatting(self, output_path: str):
        """Apply additional formatting to the Excel file"""
        try:
            wb = load_workbook(output_path)

            # Format main sheet
            if 'Flatrate Jobs & Catagories' in wb.sheetnames:
                ws = wb['Flatrate Jobs & Catagories']

                # Bold the column names row (row 2)
                for cell in ws[2]:
                    cell.font = Font(bold=True)

                # Auto-adjust column widths (basic)
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Format filters sheet
            if 'Filters ' in wb.sheetnames:
                ws = wb['Filters ']

                # Bold the column names row (row 2)
                for cell in ws[2]:
                    cell.font = Font(bold=True)

            wb.save(output_path)
        except Exception as e:
            print(f"Warning: Could not apply formatting: {e}")

    def convert(self, input_path: str, output_path: str = None):
        """
        Main conversion method

        Args:
            input_path: Path to silver JSON file
            output_path: Path to output Excel file (optional)
        """
        # Generate output path if not provided
        if output_path is None:
            input_file = Path(input_path)
            output_dir = Path("data/gold")
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = output_dir / f"{input_file.stem}_output.xlsx"

        print(f"Loading silver JSON: {input_path}")
        silver_data = self.load_silver_json(input_path)

        print(f"Processing {len(silver_data.get('systems', []))} systems...")
        rows = self.process_systems(silver_data)

        print(f"Creating Excel file...")
        self.create_excel(rows, str(output_path))

        return str(output_path)


def main(input_path: str, output_path: str = None, costbook_title: str = "WinSupply"):
    """
    Main entry point for CLI

    Args:
        input_path: Path to silver JSON file
        output_path: Path to output Excel file (optional)
        costbook_title: Title for the costbook
    """
    loader = SilverToExcelLoader(costbook_title=costbook_title)
    output_file = loader.convert(input_path, output_path)
    print(f"\nConversion complete!")
    print(f"Output: {output_file}")
    return output_file


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python -m src.stage3_loader.silver_to_excel_loader <input_json> [output_xlsx] [costbook_title]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    costbook_title = sys.argv[3] if len(sys.argv) > 3 else "WinSupply"

    main(input_path, output_path, costbook_title)
